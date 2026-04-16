"""
gdrive.py — Google Drive 上傳客戶資訊
在「客戶資料」資料夾內自動建立子資料夾（代號_品牌名），並上傳客戶基本資料 Excel 檔
"""
import io
import os
import logging
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload
import openpyxl

from parser import QuoteData

logger = logging.getLogger(__name__)

# 客戶資料根資料夾 ID
CUSTOMER_FOLDER_ID = '1ATGnrrpUHr6K9fLAIw-avPzP9-1GDVlN'


def _get_drive_service():
    """Build Google Drive service using OAuth refresh token."""
    client_id = os.environ.get('GOOGLE_CLIENT_ID')
    client_secret = os.environ.get('GOOGLE_CLIENT_SECRET')
    refresh_token = os.environ.get('GOOGLE_REFRESH_TOKEN')

    if not all([client_id, client_secret, refresh_token]):
        raise RuntimeError('Missing Google OAuth environment variables')

    creds = Credentials(
        token=None,
        refresh_token=refresh_token,
        client_id=client_id,
        client_secret=client_secret,
        token_uri='https://oauth2.googleapis.com/token',
    )
    return build('drive', 'v3', credentials=creds)


def _find_folder(service, parent_id: str, folder_name: str) -> str | None:
    """Find a folder by name under parent. Returns folder ID or None."""
    query = (
        f"'{parent_id}' in parents "
        f"and name = '{folder_name}' "
        f"and mimeType = 'application/vnd.google-apps.folder' "
        f"and trashed = false"
    )
    results = service.files().list(q=query, fields='files(id, name)', pageSize=1).execute()
    files = results.get('files', [])
    return files[0]['id'] if files else None


def _create_folder(service, parent_id: str, folder_name: str) -> str:
    """Create a folder under parent. Returns new folder ID."""
    metadata = {
        'name': folder_name,
        'mimeType': 'application/vnd.google-apps.folder',
        'parents': [parent_id],
    }
    folder = service.files().create(body=metadata, fields='id').execute()
    return folder['id']


def _create_customer_excel(data: QuoteData) -> bytes:
    """Create an Excel file with customer basic info."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '客戶基本資料'

    # Header style
    from openpyxl.styles import Font, Alignment, Border, Side
    bold_font = Font(bold=True, size=12)
    normal_font = Font(size=11)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    # Title
    ws.merge_cells('A1:B1')
    ws['A1'] = '客戶基本資料'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 45

    # Data rows
    info = data.info
    rows = [
        ('品牌名稱', info.brand_name),
        ('代號', info.code),
        ('公司抬頭', info.title),
        ('聯絡人', info.contact_name),
        ('聯絡電話', info.phone),
        ('店鋪地址', info.address),
        ('發票地址', info.invoice_address),
        ('統一編號', info.tax_id),
        ('報價日期', info.date),
        ('報價類型', data.doc_title),
    ]

    for i, (label, value) in enumerate(rows, start=3):
        cell_a = ws.cell(row=i, column=1, value=label)
        cell_a.font = bold_font
        cell_a.border = thin_border
        cell_a.alignment = Alignment(vertical='center')

        cell_b = ws.cell(row=i, column=2, value=value or '')
        cell_b.font = normal_font
        cell_b.border = thin_border
        cell_b.alignment = Alignment(vertical='center')

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def upload_customer_info(data: QuoteData) -> str:
    """
    Upload customer info Excel to Google Drive.
    Creates folder: 客戶資料/{代號}_{品牌名}/客戶基本資料.xlsx
    Returns the folder name created.
    """
    info = data.info

    # Build folder name: 代號_品牌名
    parts = []
    if info.code:
        parts.append(info.code)
    if info.brand_name:
        parts.append(info.brand_name)
    if not parts:
        parts.append(info.title or '未命名客戶')
    folder_name = '_'.join(parts)

    service = _get_drive_service()

    # Find or create customer subfolder
    folder_id = _find_folder(service, CUSTOMER_FOLDER_ID, folder_name)
    if not folder_id:
        folder_id = _create_folder(service, CUSTOMER_FOLDER_ID, folder_name)
        logger.info(f'Created folder: {folder_name}')
    else:
        logger.info(f'Folder already exists: {folder_name}')

    # Create and upload Excel
    excel_bytes = _create_customer_excel(data)
    file_metadata = {
        'name': '客戶基本資料.xlsx',
        'parents': [folder_id],
        'mimeType': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    }

    # Check if file already exists (update instead of creating duplicate)
    existing = _find_file(service, folder_id, '客戶基本資料.xlsx')
    media = MediaIoBaseUpload(
        io.BytesIO(excel_bytes),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )

    if existing:
        service.files().update(fileId=existing, media_body=media).execute()
        logger.info(f'Updated: {folder_name}/客戶基本資料.xlsx')
    else:
        service.files().create(body=file_metadata, media_body=media, fields='id').execute()
        logger.info(f'Uploaded: {folder_name}/客戶基本資料.xlsx')

    return folder_name


def _find_file(service, parent_id: str, file_name: str) -> str | None:
    """Find a file by name under parent. Returns file ID or None."""
    query = (
        f"'{parent_id}' in parents "
        f"and name = '{file_name}' "
        f"and trashed = false"
    )
    results = service.files().list(q=query, fields='files(id)', pageSize=1).execute()
    files = results.get('files', [])
    return files[0]['id'] if files else None
